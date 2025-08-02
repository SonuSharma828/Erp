from django.shortcuts import render
from employees.models import Employee
from inventory.models import Inventory
from projects.models import Project
from django.db.models import Q
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required
def report_page(request):
    # Get all data for filters
    employees = Employee.objects.all()
    inventory = Inventory.objects.all()
    projects = Project.objects.all()

    # Get filtered values from form
    employee_filter = request.GET.get('employee', '')
    inventory_filter = request.GET.get('inventory', '')
    project_filter = request.GET.get('project', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Start with all projects
    reports = Project.objects.all()

    # Apply Project Filter
    if project_filter:
        reports = reports.filter(id=project_filter)

    # Apply Date Range Filter (if both dates provided)
    if start_date and end_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        reports = reports.filter(startdate__gte=start_date_obj, deadline__lte=end_date_obj)

    # Context for template
    context = {
        'employees': employees,
        'inventory': inventory,
        'projects': projects,
        'reports': reports,
    }

    return render(request, 'report.html', context)

@login_required
def export_report_pdf(request):
    # Fetch reports based on filters (this can be customized)
    employee_filter = request.GET.get('employee', '')
    inventory_filter = request.GET.get('inventory', '')
    project_filter = request.GET.get('project', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Filter data based on parameters (you may need to adjust this as per your actual data model)
    reports = Employee.objects.all()

    if start_date and end_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        reports = reports.filter(date__range=[start_date_obj, end_date_obj])

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    # Create PDF document using reportlab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # PDF content: Simple table headers and data
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 50, "Report for Employees")
    
    y_position = height - 80
    p.drawString(50, y_position, "ID")
    p.drawString(150, y_position, "Name")
    p.drawString(250, y_position, "Details")
    p.drawString(350, y_position, "Date")

    y_position -= 20

    # Add each report row
    for report in reports:
        p.drawString(50, y_position, str(report.id))
        p.drawString(150, y_position, report.name)
        p.drawString(250, y_position, report.details)
        p.drawString(350, y_position, str(report.date))
        y_position -= 20

    p.showPage()
    p.save()

    return response

@login_required
def export_report_excel(request):
    # Fetch reports based on filters (this can be customized)
    employee_filter = request.GET.get('employee', '')
    inventory_filter = request.GET.get('inventory', '')
    project_filter = request.GET.get('project', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Filter data based on parameters (adjust as needed)
    reports = Employee.objects.all()

    if start_date and end_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        reports = reports.filter(date__range=[start_date_obj, end_date_obj])

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Report"

    # Add headers
    ws['A1'] = 'ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Details'
    ws['D1'] = 'Date'

    # Populate rows with report data
    row_num = 2
    for report in reports:
        ws[f'A{row_num}'] = report.id
        ws[f'B{row_num}'] = report.name
        ws[f'C{row_num}'] = report.details
        ws[f'D{row_num}'] = str(report.date)
        row_num += 1

    # Create response to send the Excel file to the browser
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'

    wb.save(response)
    return response

